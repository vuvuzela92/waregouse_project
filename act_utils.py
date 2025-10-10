import pandas as pd
import json
import base64
from datetime import datetime, timedelta
import io
import asyncio
import aiohttp
import os
import logging
import asyncpg
from dotenv import load_dotenv
import zipfile
from openpyxl import load_workbook


def batchify(data, batch_size):
    """
    Splits data into batches of a specified size.

    Parameters:
    - data: The list of items to be batched.
    - batch_size: The size of each batch.

    Returns:
    - A generator yielding batches of data.
    """
    for i in range(0, len(data), batch_size):
        yield data[i:i + batch_size]

# Функция для загрузки API токенов из файла tokens.json
def load_api_tokens():
    # Укажи полный путь к tokens.json
    # Определяем директорию текущего скрипта
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(CURRENT_DIR, "tokens.json")
    # file_path = 'tokens.json'
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл с токенами не найден: {file_path}")
    
    with open(file_path, encoding='utf-8') as f:
        tokens = json.load(f)
        return tokens
    

semaphore = asyncio.Semaphore(10)

async def documents_list_async(token, title, days_back=10):
    """Получить все документы с пагинацией"""
    url = 'https://documents-api.wildberries.ru/api/v1/documents/list'
    headers = {'Authorization': token}
    beginTime = (datetime.now()-timedelta(days=days_back)).strftime('%Y-%m-%d')
    endTime = datetime.now().strftime('%Y-%m-%d')
    
    all_documents = []
    offset = 0
    limit = 50
    
    async with semaphore:
        async with aiohttp.ClientSession(headers=headers) as session:
            while True:
                params = {
                    'beginTime': beginTime,
                    'endTime': endTime,
                    'category': title,
                    'limit': limit,
                    'offset': offset
                }
                
                try:
                    async with session.get(url, params=params) as res:
                        if res.status == 200:
                            data = await res.json()
                            documents = data['data']['documents']
                            
                            # Если документов нет или получено меньше limit - это последняя страница
                            if not documents or len(documents) < limit:
                                if documents:  # добавляем последние документы
                                    all_documents.extend(documents)
                                    print(f"Последняя страница: получено {len(all_documents)} документов")
                                break
                                
                            all_documents.extend(all_documents)
                            print(f"Получено {len(all_documents)} документов, offset: {offset}")
                            offset += limit
                            await asyncio.sleep(10)
                        else:
                            print(f"Ошибка {res.status}")
                            await asyncio.sleep(10)
                            break
                            
                except aiohttp.ClientError as err:
                    print(f"Сетевая ошибка: {err}")
                    await asyncio.sleep(10)
                    break
                except Exception as e:
                    print(f"Неожиданная ошибка: {e}")
                    await asyncio.sleep(10)
                    break
        
        if all_documents:
            df = pd.DataFrame(all_documents)
            print(f"Всего получено {len(df)} документов за период {beginTime} - {endTime}")
            return df
        else:
            print("Не удалось получить документы")
            return None
        

# Глобальный лок для создания таблицы
table_creation_lock = asyncio.Lock()
async def create_insert_table_db_async(df: pd.DataFrame, table_name: str, columns_type: dict, key_columns: tuple):
    load_dotenv()

    conn = None
    try:
        conn = await asyncpg.connect(
            user=os.getenv('USER_2'),
            password=os.getenv('PASSWORD_2'),
            database=os.getenv('NAME_2'),
            host=os.getenv('HOST_2'),
            port=os.getenv('PORT_2'),
            timeout=10
        )

        # Проверка типов данных
        valid_types = ['INTEGER', 'BIGINT', 'SMALLINT', 'NUMERIC', 'DATE', 'TIMESTAMP', 'BOOLEAN', 'TEXT', 'VARCHAR']
        for col, dtype in columns_type.items():
            if not dtype.strip():
                raise ValueError(f"Пустой тип данных для колонки {col}")
            base_type = dtype.split('(')[0].strip().upper()
            if base_type not in valid_types:
                raise ValueError(f"Недопустимый тип данных для колонки {col}: {dtype}")
            if '(' in dtype and base_type not in ['NUMERIC', 'VARCHAR']:
                raise ValueError(f"Скобки недопустимы для типа {base_type} в колонке {col}: {dtype}")

        # Проверка и добавление отсутствующих колонок
        missing_cols = set(columns_type.keys()) - set(df.columns)
        for col in missing_cols:
            df[col] = None
            logging.info(f"Добавлена отсутствующая колонка {col} с None")
        extra_cols = set(df.columns) - set(columns_type.keys())
        if extra_cols:
            logging.warning(f"Лишние колонки в DataFrame: {extra_cols}, они будут проигнорированы")

        # Формирование SQL для колонок
        columns_definition = ", ".join([f"{col} {dtype}" for col, dtype in columns_type.items()])
        # создает SQL-выражение для уникального ограничения (UNIQUE constraint) в таблице базы данных.
        unique_constraint = f"CONSTRAINT unique_{table_name} UNIQUE ({', '.join(key_columns)})" if key_columns else ""

        # Отладка: вывод SQL-запроса
        create_table_query = f"""
            DO $$
            BEGIN
                IF NOT EXISTS (SELECT FROM pg_tables WHERE tablename = '{table_name}') THEN
                    CREATE TABLE {table_name} (
                        {columns_definition}{', ' + unique_constraint if unique_constraint else ''}
                    );
                END IF;
            END$$;
        """



        # Безопасное создание таблицы
        async with table_creation_lock:
            await conn.execute(create_table_query)
        
        # Подготовка данных для вставки
        columns = list(columns_type.keys())  # Используем только колонки из columns_type
        records = [tuple(None if pd.isna(x) else x for x in row) for row in df[columns].to_records(index=False)]

        # Формирование UPSERT-запроса
        updates = ', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col not in key_columns])
        query = f"""
            INSERT INTO {table_name} ({', '.join(columns)})
            VALUES ({', '.join([f'${i+1}' for i in range(len(columns))])})
            ON CONFLICT ON CONSTRAINT unique_{table_name}
            DO UPDATE SET {updates}
        """
        
        await conn.executemany(query, records)
        logging.info(f"Успешно сохранено {len(df)} строк в {table_name}")
        
    except Exception as e:
        logging.error(f"Ошибка при работе с БД: {str(e)}")
        raise
    finally:
        if conn:
            await conn.close()


# Получаем перечень доступных для скачивания документов
async def create_acceptance_certificate_marketplace_async():
    """Функция обрабатывает данные полученные в documents_list
    для каждого кабинета на ВБ. Забирает информацию об АПП ФБОи ФБС"""
    tokens = load_api_tokens()
    
    doc_list_data = []
    
    for account, token in tokens.items():
        # Получаем документы для 'act-income-mp'
        result_mp = await documents_list_async(token, 'act-income-mp')
        if result_mp is not None and not result_mp.empty:
            result_mp['account'] = account
            result_mp['doc_type'] = 'act-income-mp'
            doc_list_data.append(result_mp)
            print(f"Аккаунт {account}: получено {len(result_mp)} документов act-income-mp")
        else:
            print(f"Аккаунт {account}: нет документов act-income-mp")

        # Получаем документы для 'act-income'
        result_income = await documents_list_async(token, 'act-income')
        if result_income is not None and not result_income.empty:  
            result_income['account'] = account
            result_income['doc_type'] = 'act-income'
            doc_list_data.append(result_income)
            print(f"Аккаунт {account}: получено {len(result_income)} документов act-income")
        else:
            print(f"Аккаунт {account}: нет документов act-income")

    # Объединяем все данные, если они есть
    if doc_list_data:
        documents_list_df = pd.concat(doc_list_data, ignore_index=True)
        print(f"Всего собрано {len(documents_list_df)} документов")
        return documents_list_df
    else:
        print("Не удалось получить ни одного документа")
        return pd.DataFrame()  # возвращаем пустой DataFrame вместо None
    
async def get_decoded_acts_fbs_async(account: str, doc_list: list, tokens):
    """Асинхронная версия: позволяет обрабатывать архивный файл,
    полученный из метода create_acceptance_certificate_marketplace
    
    Учитывает ограничения API:
    - Лимит: 1 запрос в 5 минут на аккаунт
    - Максимум 50 документов в одном запросе
    """
    
    with open('tokens.json', 'r', encoding='utf-8') as file:
        tokens = json.load(file)

    url = 'https://documents-api.wildberries.ru/api/v1/documents/download/all'
    headers = {'Authorization': tokens[account]}

    # Проверяем лимит в 50 документов
    if len(doc_list) > 50:
        print(f"Предупреждение: запрошено {len(doc_list)} документов, но максимум 50. Будут обработаны первые 50.")
        doc_list = doc_list[:50]

    payload = {
        "params": [
            {
                "extension": "xlsx",  # согласно документации - формат документа
                "serviceName": doc_id  # уникальный ID документа
            } for doc_id in doc_list
        ]
    }
    
    async with aiohttp.ClientSession(
        headers=headers,
        timeout=aiohttp.ClientTimeout(total=300)  # 5 минут таймаут
    ) as session:
        async with semaphore:
            try:
                async with session.post(url, json=payload) as res:
                    if res.status == 200:
                        data = await res.json()
                        
                        # Проверяем структуру ответа согласно документации
                        if 'data' in data and 'document' in data['data']:
                            document_data = data['data']['document']
                            
                            # Дополнительные поля из ответа (опционально)
                            file_name = data['data'].get('fileName', 'unknown')
                            extension = data['data'].get('extension', 'xlsx')
                            
                            print(f"Успешно получен документ: {file_name}.{extension} для аккаунта {account}")
                            
                            # Декодируем base64
                            try:
                                decoded_data = base64.b64decode(document_data)
                                decoded_acts = io.BytesIO(decoded_data)
                                return decoded_acts
                            except base64.binascii.Error as e:
                                print(f"Ошибка декодирования base64 для аккаунта {account}: {e}")
                                return None
                                
                        else:
                            print(f"Неожиданная структура ответа для аккаунта {account}")
                            return None
                            
                    elif res.status == 400:
                        error_data = await res.json()
                        error_detail = error_data.get('detail', 'Неправильный запрос')
                        print(f"Ошибка 400 для аккаунта {account}: {error_detail}")
                        return None
                        
                    elif res.status == 401:
                        print(f"Ошибка 401 для аккаунта {account}: Не авторизован")
                        return None
                        
                    elif res.status == 429:
                        error_data = await res.json()
                        error_detail = error_data.get('detail', 'Слишком много запросов')
                        print(f"Ошибка 429 для аккаунта {account}: {error_detail}")
                        print("Лимит: 1 запрос в 5 минут на аккаунт")
                        # Ждем перед повторной попыткой
                        await asyncio.sleep(300)  # 5 минут
                        return None
                        
                    else:
                        print(f"Неизвестная ошибка {res.status} для аккаунта {account}")
                        return None
                        
            except aiohttp.ClientError as e:
                print(f"Сетевая ошибка для аккаунта {account}: {e}")
                return None
                
            except asyncio.TimeoutError:
                print(f"Таймаут запроса для аккаунта {account}")
                return None
                
            except Exception as e:
                print(f'Неожиданная ошибка для аккаунта {account}: {e}')
                return None
            
async def get_decoded_acts_fbs(account, doc_list, tokens):
    batch_doc_list = list(batchify(doc_list, 50))
    url = 'https://documents-api.wildberries.ru/api/v1/documents/download/all'
    headers = {'Authorization': tokens[account]}
    acts_data = []
    async with semaphore:
        async with aiohttp.ClientSession(headers=headers) as session:
            for batch in batch_doc_list:
                payload = {
                            "params": [
                                {
                                    "extension": "xlsx",
                                    "serviceName": doc_id
                                } for doc_id in batch
                            ]
                        }
                print(payload)
                try:
                    async with session.post(url, json=payload) as res:
                        print(res.status)
                        error = await res.json()
                        if res.status == 400:
                            print(f"Ошибка 400 {account}: {error.get('message') or error}")
                        if res.status == 429:
                            print("429 ошибка — лимит запросов. Ждём 5 минут")
                            await asyncio.sleep(300)
                            continue
                        if res.status == 401:
                            print(f"401 ошибка авторизации по ЛК {account}")
                            await asyncio.sleep(300)
                            continue
                        if res.status == 200:
                            data = await res.json()
                            document_data = data['data']['document']
                            # Извлекаем зипы документов
                            decoded_data = base64.b64decode(document_data)
                            # Получаю общий архив, со всеми запрошенными документами
                            decoded_acts = io.BytesIO(decoded_data)
                            # Обрабатываем полученные документы
                            acts_data.append(decoded_acts)
                            return acts_data
                        else:
                            print('Отсутствует список документов')
                except aiohttp.ClientError as e:
                    print(f"Сетевая ошибка для {account}: {e}")
                    await asyncio.sleep(30)


def proccessing_data_acceptance_act(account, decoded_acts):
    """Позволяет обрабатывать каждый отдельный архив, который
    содержит внутри эксель файл с информацией о приеме передаче товара
    по заказам ФБС. Возвращает датафрейм с информацией по каждому
    сборочному заданию в документе."""
    acts_list = []
    # Открываем данные общего архива со всеми документами
    with zipfile.ZipFile(decoded_acts, 'r') as full_zip:
        # Представляем документы в архиве, как коллекцию zip-файлов
        acts_income_zip = full_zip.namelist()
        for act in acts_income_zip:
            # Проходим циклом по каждому документу
            with full_zip.open(act) as nested_zip_file:
                # Читаем содержимое вложенного ZIP как байты
                nested_zip_bytes = nested_zip_file.read()
                # Загружаем его в память как новый ZIP-объект. Внутри как правило два файла, один эксель, другой .sig
                nested_zip = zipfile.ZipFile(io.BytesIO(nested_zip_bytes))
                # Получаем список файлов внутри вложенного ZIP
                print("Файлы внутри вложенного ZIP:")
                # Теперь извлекаем каждый эксель файл
                for inner_file in nested_zip.namelist():
                    print(f"  - {inner_file}")
                    # Если нужно, обрабатываем, например, ищем .xlsx
                    if inner_file.endswith('.xlsx'):
                        with nested_zip.open(inner_file) as xlsx_file:
                            wb = load_workbook(xlsx_file, read_only=True)
                            ws = wb.active
                            # Получаем значение из D3
                            if ws['D3'].value:
                                date_value = ws['D3'].value
                            elif ws['F3'].value:
                                date_value = ws['F3'].value
                            else:
                                date_value = 'Нет даты'
                            df_act = pd.read_excel(xlsx_file, header=None)

                            # Получаем строки 8 и 9 как заголовки
                            header_1 = df_act.iloc[8].fillna('')
                            header_2 = df_act.iloc[9].fillna('')

                            # Формируем комбинированные заголовки
                            multi_header = [
                                f"{h1.strip()} - {h2.strip()}" if h2.strip() else h1.strip()
                                for h1, h2 in zip(header_1, header_2)
                            ]

                            print(f"Количество колонок в df: {df_act.shape[1]}")
                            print(f"Количество заголовков: {len(multi_header)}")
                            
                            # Загружаем данные с 12 строки
                            df = df_act.iloc[12:].copy()
                            df.columns = multi_header
                            df = df.dropna(how='all').reset_index(drop=True)
                            df['Документ'] = act
                            df['Дата'] = date_value
                            df['Дата'] = df['Дата'].str.replace(' г.', '').str.strip()
                            df['Номер_документа'] = df['Документ'].str.extract(r'act-income-mp-(\d+)\.zip')        
                            df['account'] = account
                            acts_list.append(df)
                            print(f'Данные по {act} добавлены в список')
    return acts_list                    

async def get_all_fbs_acts_async():
    """Получает и обрабатывает акты ПП ФБС/ФБО со всех ЛК"""
    # Собирает полный перечень актов
    full_docs = []
    
    # ДОБАВЛЕНО AWAIT - получаем DataFrame с документами
    acceptance_certificate_df = await create_acceptance_certificate_marketplace_async()
    
    # Проверяем, что DataFrame не пустой
    if acceptance_certificate_df.empty:
        print("Нет документов для обработки")
        return pd.DataFrame(), pd.DataFrame()
    
    # Группируем по аккаунту и получаем список документов для каждого аккаунта
    acceptance_certificate_dict = acceptance_certificate_df.groupby('account')['serviceName'].apply(list).to_dict()
    
    # Для каждого аккаунта получаем документы, которые к нему относятся
    for account, doc_list in acceptance_certificate_dict.items():
        print(f"Обрабатываем аккаунт {account}, документов: {len(doc_list)}")
        
        # Получаем распакованные из zip акты
        acts = await get_decoded_acts_fbs(account, doc_list, tokens=load_api_tokens())
        
        if acts and acts[0]:
            # Обрабатываем данные актов
            processed_acts = proccessing_data_acceptance_act(account, acts[0])
            full_docs.extend(processed_acts)
        else:
            print(f"Не удалось получить документы для аккаунта {account}")
    
    # Если нет обработанных документов, возвращаем пустые DataFrame
    if not full_docs:
        print("Нет обработанных документов")
        return pd.DataFrame(), pd.DataFrame()
    
    # Объединяем датафреймы
    final_df = pd.concat(full_docs, ignore_index=True)
    
    # Проверяем, что final_df не пустой
    if final_df.empty:
        print("Объединенный DataFrame пуст")
        return pd.DataFrame(), pd.DataFrame()
    
    # Удаляем лишние данные из объединенного датафрейма 
    final_df = final_df[final_df['Фактически принято - Стикер/этикетка'] != 'Итого']
    
    # Из полученных данных формируем акты-приема передачи для ФБО
    fbo_acts_df = final_df[['№ п\п', 'Товар (наименование)', 'Ед. изм.', 'Фактически принято - баркод', ' - артикул продавца', ' - сорт, размер', ' - КИЗ', ' - ШК короба', ' - кол-во', 'Документ','Номер_документа', 'Дата', 'account']]
    
    # ВБ иногда путает акты ФБО и ФБС, поэтому фильтруем по ШК короба
    fbo_acts_df = fbo_acts_df[fbo_acts_df[' - ШК короба'].notna()]
    
    # Приводим названия колонок к читаемому виду
    fbo_acts_df = fbo_acts_df.rename(columns={
        '№ п\\п': 'num',
        'Товар (наименование)': 'product_name',
        'Ед. изм.': 'unit',
        'Фактически принято - баркод': 'barcode',
        ' - артикул продавца': 'vendor_code',
        ' - сорт, размер': 'size',
        ' - КИЗ': 'kiz',
        ' - ШК короба': 'box_barcode',
        ' - кол-во': 'quantity',
        'Документ': 'document',
        'Номер_документа': 'document_number',
        'Дата': 'date',
        'account': 'account'
    })
    
    # Заменяем пустоты
    fbo_acts_df['kiz'] = fbo_acts_df['kiz'].fillna('Нет КИЗов')
    
    # Приводим колонку с датой к нужному формату
    fbo_acts_df['date'] = fbo_acts_df['date'].str.replace('"','').str.replace(' ', '').str.replace('г.', '')
    fbo_acts_df['date'] = pd.to_datetime(fbo_acts_df['date'], format='%d%m%Y', errors='coerce')
    
    # Удаляем лишние символы из номера документа
    fbo_acts_df['document_number'] = fbo_acts_df['document'].str.extract(r'(\d+)\.zip')[0]
    print('Данные по ФБО получены')

    # Из полученных данных формируем акты-приема передачи для ФБС
    fbs_acts_df = final_df[['№ п\п', 'Номер заказа', 'Ед. изм.', 'Фактически принято - Стикер/этикетка', ' - Кол-во', 'Документ','Номер_документа', 'Дата', 'account']]
    
    # ВБ иногда путает акты ФБО и ФБС, поэтому фильтруем по стикерам
    fbs_acts_df = fbs_acts_df[fbs_acts_df['Фактически принято - Стикер/этикетка'].notna()]
    
    # Приводим названия колонок к читаемому виду
    fbs_acts_df = fbs_acts_df.rename(columns={
        '№ п\\п': 'num',
        'Номер заказа': 'order_number',
        'Ед. изм.': 'unit',
        'Фактически принято - Стикер/этикетка': 'sticker',
        ' - Кол-во': 'quantity',
        'Документ': 'document',
        'Номер_документа': 'document_number',
        'Дата': 'date',
        'account': 'account'
    })
    
    return fbs_acts_df, fbo_acts_df

async def main():
    """Исполняемая функция по получению актов ПП с ВБ"""
    # Получаем данные по актам ФБС/ФБО
    fbs_df, fbo_df = await get_all_fbs_acts_async()

    # Обработка FBO данных
    if not fbo_df.empty:
        fbo_df = fbo_df.where(pd.notnull(fbo_df), None)  # NaN → None
        # Исправляем преобразование даты для FBO
        fbo_df['date'] = pd.to_datetime(fbo_df["date"], dayfirst=True, errors='coerce').dt.date
        fbo_df['num'] = fbo_df['num'].astype(int)
        fbo_df['quantity'] = fbo_df['quantity'].astype(int)
        
        columns_type_fbo = {
            'num': 'INTEGER',
            'product_name': 'VARCHAR(255)',
            'unit': 'VARCHAR(50)',
            'barcode': 'VARCHAR(50)', 
            'vendor_code': 'VARCHAR(50)',
            'size': 'VARCHAR(50)',
            'kiz': 'VARCHAR(255)',
            'box_barcode': 'VARCHAR(50)',
            'quantity': 'INTEGER',
            'document': 'VARCHAR(255)',
            'document_number': 'VARCHAR(50)',
            'date': 'DATE',
            'account': 'VARCHAR(100)'
        }

        key_cols_fbo = ('vendor_code', 'box_barcode', 'document_number')
        table_name_fbo = 'acceptance_fbo_acts_new'
        await create_insert_table_db_async(fbo_df, table_name_fbo, columns_type_fbo, key_cols_fbo)
    else:
        print("Нет данных FBO для вставки в БД")

    # Обработка FBS данных
    if not fbs_df.empty:
        fbs_df = fbs_df.astype(object)
        fbs_df = fbs_df.where(pd.notnull(fbs_df), None)  # NaN → None
        
        # ИСПРАВЛЕНО: правильное преобразование даты с dayfirst=True
        fbs_df['date'] = pd.to_datetime(fbs_df["date"], dayfirst=True, errors='coerce').dt.date
        
        # ЯВНОЕ приведение типов для КАЖДОЙ колонки
        fbs_df['num'] = pd.to_numeric(fbs_df['num'], errors='coerce').fillna(0).astype(int)
        fbs_df['order_number'] = fbs_df['order_number'].astype(str)
        fbs_df['unit'] = fbs_df['unit'].astype(str)
        fbs_df['sticker'] = fbs_df['sticker'].astype(str)
        fbs_df['quantity'] = pd.to_numeric(fbs_df['quantity'], errors='coerce').fillna(0).astype(int)
        fbs_df['document'] = fbs_df['document'].astype(str)
        fbs_df['document_number'] = fbs_df['document_number'].astype(str)
        fbs_df['account'] = fbs_df['account'].astype(str)

        columns_type_fbs = {
            'num': 'INTEGER',
            'order_number': 'VARCHAR(255)',
            'unit': 'VARCHAR(50)',
            'sticker': 'VARCHAR(255)',
            'quantity': 'INTEGER', 
            'document': 'VARCHAR(255)',
            'document_number': 'BIGINT',
            'date': 'DATE',
            'account': 'VARCHAR(50)', 
        }

        key_cols_fbs = ('order_number', 'sticker', 'document_number')
        table_name_fbs = 'acceptance_fbs_acts_new'
        await create_insert_table_db_async(fbs_df, table_name_fbs, columns_type_fbs, key_cols_fbs)
    else:
        print("Нет данных FBS для вставки в БД")