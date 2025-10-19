import requests
import pandas as pd
import json
from datetime import datetime, timedelta
import base64
import io
import zipfile
from openpyxl import load_workbook
import re
import os
import logging
import asyncpg
from dotenv import load_dotenv
import asyncio
import aiohttp
from sqlalchemy import create_engine, text, DDL
from sqlalchemy.exc import SQLAlchemyError
import logging
import os
from dotenv import load_dotenv
import psycopg2
from psycopg2 import OperationalError

# Функция для загрузки API токенов из файла tokens.json
def load_api_tokens():
    # Укажи полный путь к tokens.json
    # Определяем директорию текущего скрипта
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(CURRENT_DIR, "tokens.json")
    # file_path = 'tokens.json'
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл с токенами не найден: {file_path}")
    
    with open(file_path, encoding='utf-8') as f:
        tokens = json.load(f)
        return tokens

def doc_categories():
    """Функция реализует получение категорий документов, 
    по методу 
    https://dev.wildberries.ru/openapi/financial-reports-and-accounting#tag/Dokumenty/paths/~1api~1v1~1documents~1categories/get
    Лимит запросов на один аккаунт продавца:
    Период	    Лимит	    Интервал	Всплеск
    10 секунд	1 запрос    10 секунд   5 запросов """
    # Загружаем токены
    with open('tokens.json', 'r', encoding='utf-8') as file:
        tokens = json.load(file)

    # Определяем адрес запроса для получения документов 
    url = 'https://documents-api.wildberries.ru/api/v1/documents/categories'
    headers = {'Authorization' : tokens['Вектор']}
    # Отправляе get запрос для получения категорий документов
    try:
        res = requests.get(url, headers=headers, timeout=2)
        if res.status_code == 200:
            doc_categories = pd.DataFrame(res.json()['data']['categories'])
            return doc_categories
    # для отлавливания любых ошибок, которые могут возникнуть при выполнении запроса
    except requests.exceptions.RequestException as e:
        print(f'"Ошибка запроса:", {e}')
        return None
    # для отлавливания других ошибок
    except Exception as e:
        print(f'Ошибка не связанная с запросом {e}')
        return None


# beginTime = (datetime.now()-timedelta(days=10)).strftime('%Y-%m-%d')
# endTime = (datetime.now()-timedelta(days=1)).strftime('%Y-%m-%d')   

def documents_list(token, title):
    """Метод предоставляет список документов продавца. 
    Вы можете получить один или несколько документов из полученного списка."""
    
    url = 'https://documents-api.wildberries.ru/api/v1/documents/list'
    headers = {'Authorization' : token}
    beginTime = (datetime.now()-timedelta(days=200)).strftime('%Y-%m-%d')
    endTime = (datetime.now()-timedelta(days=0)).strftime('%Y-%m-%d') 
    params = {'beginTime' : beginTime,
              'endTime' : endTime,
              'category': title}
    
    try:
        res = requests.get(url, headers=headers, params=params)
        if res.status_code == 200:
            data = res.json()
            df = pd.DataFrame(data['data']['documents'])
            return df
        else:
            print('Отсутствует список документов')
    except requests.exceptions.RequestException as e:
        print(f'"Ошибка запроса:", {e}')
        return None
    # для отлавливания других ошибок
    except Exception as e:
        print(f'Ошибка не связанная с запросом {e}')
        return None
    
# Получаем перечень доступных для скачивания документов
def create_acceptance_act():
    """Функция обрабатывает данные полученные в documents_list
    для каждого кабинета на ВБ"""
    
    with open('tokens.json', 'r', encoding='utf-8') as file:
        tokens = json.load(file)

    doc_list_data = []
    for account, token in tokens.items():
        # Получаем список документов для ФБО
        df = pd.DataFrame(documents_list(token, 'act-income'))
        df['account'] = account
        doc_list_data.append(df)
        # Получаем список документов для ФБC
        df = pd.DataFrame(documents_list(token, 'act-income-mp'))
        df['account'] = account
        doc_list_data.append(df)

    documents_list_df = pd.concat(doc_list_data)
    return documents_list_df

def act_income_docs_list(download_all_acts, account):
    """Функция позволяет обрабатывать акты-приема передачи, 
        полученные от ВБ при приеме товара для продаж по системе ФБО"""
    acts_list = []
    # Первый уровень — внешний архив
    zip_data = io.BytesIO(download_all_acts)
    with zipfile.ZipFile(zip_data, 'r') as full_zip:
        acts_income_zip = full_zip.namelist()
        for act in acts_income_zip:
            with full_zip.open(act) as nested_zip_file:
                # Читаем содержимое вложенного ZIP как байты
                nested_zip_bytes = nested_zip_file.read()
                # Загружаем его в память как новый ZIP-объект
                nested_zip = zipfile.ZipFile(io.BytesIO(nested_zip_bytes))
                # Получаем список файлов внутри вложенного ZIP
                print("Файлы внутри вложенного ZIP:")
                for inner_file in nested_zip.namelist():
                    print(f"  - {inner_file}")
                    # Если нужно, обрабатываем, например, ищем .xlsx
                    if inner_file.endswith('.xlsx'):
                        with nested_zip.open(inner_file) as xlsx_file:
                            # df = pd.read_excel(xlsx_file)
                            # print(f"Первые строки Excel-файла {inner_file}:")
                            # print(df.head())
                            df_act = pd.read_excel(xlsx_file, header=None)
                            # Получаем строки 8 и 9 как заголовки
                            header_1 = df_act.iloc[8].fillna('')
                            header_2 = df_act.iloc[9].fillna('')

                            # Формируем комбинированные заголовки
                            multi_header = [
                                f"{h1.strip()} - {h2.strip()}" if h2.strip() else h1.strip()
                                for h1, h2 in zip(header_1, header_2)
                            ]

                            print(f"Количество колонок в df: {df_act.shape[1]}")
                            print(f"Количество заголовков: {len(multi_header)}")
                            
                            # Загружаем данные с 12 строки
                            df = df_act.iloc[12:].copy()
                            df.columns = multi_header
                            df = df.dropna(how='all').reset_index(drop=True)
                            df['Документ'] = act
                            # df['Номер_документа'] = df['Документ'].str.extract(r'act-income-(\d+)\.zip')
                            df['Номер_документа'] = df['Документ'].str.extract(r'(\d+)\.zip')[0]        
                            df['account'] = account
                            acts_list.append(df)
                            print(f'Данные по {act} добавлены в список')
    return acts_list


def download_all_acts(payload, account):
    """Загружает все акты по ФБО"""
    with open('tokens.json', 'r', encoding='utf-8') as file:
        tokens = json.load(file)
   
    url = 'https://documents-api.wildberries.ru/api/v1/documents/download/all'
    headers = {'Authorization': tokens[account]}

    try:
        res = requests.post(url, json=payload, headers=headers)
        if res.status_code == 200:
            data = res.json()
            document_data = data['data']['document']
            # Извлекаем зипы документов
            decoded_data = base64.b64decode(document_data)
            # Обрабатываем полученные документы
            decoded_acts = act_income_docs_list(decoded_data, account)
            return decoded_acts
        else:
            print('Отсутствует список документов')
    except requests.exceptions.RequestException as e:
        print(f'"Ошибка запроса:", {e}')
        return None
    # для отлавливания других ошибок
    except Exception as e:
        print(f'Ошибка не связанная с запросом {e}')
        return None
    
def get_act_incomes():
    """Объединяем акты приема-передачи по ФБО
    со всех кабинетов в единный список,
    затем создаем из него единый датафрейм"""
    # Сюда помещаем все акты
    decoded_data_list = []
    act_incomes_dict = create_acceptance_act().groupby('account')['serviceName'].apply(list).to_dict()
    # Проходим циклом по каждому списку документов, для каждого аккаунта
    for account in act_incomes_dict:
        print(f'Получаем данные для {account}')
        
        payload = {
            "params": [
                {
                    "extension": "xlsx",
                    "serviceName": doc_id
                } for doc_id in act_incomes_dict[account]
            ]
        }
        print(payload)
        # Получаем и обрабатываем списки документов
        decoded_data_list.append(download_all_acts(payload, account))
    # Создаем список датафреймов для последующей обработки
    all_dfs = []
    # Складываем в список выше все датафреймы
    for doc_list in decoded_data_list:
        all_dfs.extend(doc_list)
    # Создаем итоговый датафрейм для всех кабинетов
    final_df = pd.concat(all_dfs, ignore_index=True)
    # Убираем строки, которые не отностятся к кабинету
    final_df = final_df[final_df['№ п\п'] != 'Итого']
    return final_df

# Получаем перечень доступных для скачивания документов
def create_acceptance_certificate_marketplace():
    """Функция обрабатывает данные полученные в documents_list
    для каждого кабинета на ВБ"""
    with open('tokens.json', 'r', encoding='utf-8') as file:
        tokens = json.load(file)
    doc_list_data = []
    for account, token in tokens.items():
        df = pd.DataFrame(documents_list(token, 'act-income-mp'))
        df['account'] = account
        doc_list_data.append(df)

        df = pd.DataFrame(documents_list(token, 'act-income'))
        df['account'] = account
        doc_list_data.append(df)

    documents_list_df = pd.concat(doc_list_data)
    return documents_list_df

def get_decoded_acts_fbs(account, doc_list):
    """Позволяет обрабатывать архивный файл,
        полученный из метода create_acceptance_certificate_marketplace"""
    with open('tokens.json', 'r', encoding='utf-8') as file:
        tokens = json.load(file)

    url = 'https://documents-api.wildberries.ru/api/v1/documents/download/all'
    headers = {'Authorization': tokens[account]}

    payload = {
                "params": [
                    {
                        "extension": "xlsx",
                        "serviceName": doc_id
                    } for doc_id in doc_list
                ]
            }
    try:
        res = requests.post(url, json=payload, headers=headers)
        if res.status_code == 200:
            data = res.json()
            document_data = data['data']['document']
            # Извлекаем зипы документов
            decoded_data = base64.b64decode(document_data)
            # Получаю общий архив, со всеми запрошенными документами
            decoded_acts = io.BytesIO(decoded_data)
            # Обрабатываем полученные документы
            return decoded_acts
        else:
            print('Отсутствует список документов')
    except requests.exceptions.RequestException as e:
        print(f'"Ошибка запроса:", {e}')
        return None
    # для отлавливания других ошибок
    except Exception as e:
        print(f'Ошибка не связанная с запросом {e}')
        return None
    

def proccessing_data_acceptance_act(account, decoded_acts):
    """Позволяет обрабатывать каждый отдельный архив, который
    содержит внутри эксель файл с информацией о приеме передаче товара
    по заказам ФБС. Возвращает датафрейм с информацией по каждому
    сборочному заданию в документе."""
    acts_list = []
    # Открываем данные общего архива со всеми документами
    with zipfile.ZipFile(decoded_acts, 'r') as full_zip:
        # Представляем документы в архиве, как коллекцию zip-файлов
        acts_income_zip = full_zip.namelist()
        for act in acts_income_zip:
            # Проходим циклом по каждому документу
            with full_zip.open(act) as nested_zip_file:
                # Читаем содержимое вложенного ZIP как байты
                nested_zip_bytes = nested_zip_file.read()
                # Загружаем его в память как новый ZIP-объект. Внутри как правило два файла, один эксель, другой .sig
                nested_zip = zipfile.ZipFile(io.BytesIO(nested_zip_bytes))
                # Получаем список файлов внутри вложенного ZIP
                print("Файлы внутри вложенного ZIP:")
                # Теперь извлекаем каждый эксель файл
                for inner_file in nested_zip.namelist():
                    print(f"  - {inner_file}")
                    # Если нужно, обрабатываем, например, ищем .xlsx
                    if inner_file.endswith('.xlsx'):
                        with nested_zip.open(inner_file) as xlsx_file:
                            wb = load_workbook(xlsx_file, read_only=True)
                            ws = wb.active
                            # Получаем значение из D3
                            if ws['D3'].value:
                                date_value = ws['D3'].value
                            elif ws['F3'].value:
                                date_value = ws['F3'].value
                            else:
                                date_value = 'Нет даты'
                            df_act = pd.read_excel(xlsx_file, header=None)

                            # Получаем строки 8 и 9 как заголовки
                            header_1 = df_act.iloc[8].fillna('')
                            header_2 = df_act.iloc[9].fillna('')

                            # Формируем комбинированные заголовки
                            multi_header = [
                                f"{h1.strip()} - {h2.strip()}" if h2.strip() else h1.strip()
                                for h1, h2 in zip(header_1, header_2)
                            ]

                            print(f"Количество колонок в df: {df_act.shape[1]}")
                            print(f"Количество заголовков: {len(multi_header)}")
                            
                            # Загружаем данные с 12 строки
                            df = df_act.iloc[12:].copy()
                            df.columns = multi_header
                            df = df.dropna(how='all').reset_index(drop=True)
                            df['Документ'] = act
                            df['Дата'] = date_value
                            df['Дата'] = df['Дата'].str.replace(' г.', '').str.strip()
                            df['Номер_документа'] = df['Документ'].str.extract(r'act-income-mp-(\d+)\.zip')        
                            df['account'] = account
                            acts_list.append(df)
                            print(f'Данные по {act} добавлены в список')
    return acts_list


def get_all_fbs_acts():
    """ Собирает акты со всех кабинетов, используя методы 
        get_decoded_acts_fbs и proccessing_data_acceptance_act.
        Объединяет их в единый датафрейм."""
    acceptance_certificate_marketplace_dict = create_acceptance_certificate_marketplace().groupby('account')['serviceName'].apply(list).to_dict()
    full_docs = []
    for account, doc_list in acceptance_certificate_marketplace_dict.items():
        act = get_decoded_acts_fbs(account, doc_list)
        full_docs.extend(proccessing_data_acceptance_act(account, act))
    final_df = pd.concat(full_docs)    
    final_df = final_df[final_df['Фактически принято - Стикер/этикетка'] != 'Итого']
    return final_df


def clean_and_parse_date(date_str):
    # Удаляем кавычки и букву "г"
    cleaned = re.sub(r'[\"г]', '', str(date_str)).strip()
    # Преобразуем в дату
    try:
        return datetime.strptime(cleaned, '%d %m %Y')
    except Exception:
        return None  # или np.nan, если используете numpy
    
def create_insert_table_db(df: pd.DataFrame, table_name: str, columns_type: dict, key_columns: tuple):
    """
    Загружает DataFrame в таблицу PostgreSQL с upsert по уникальному ключу.
    Если таблицы нет — создает с нужными типами и ограничением уникальности.
    """
    import psycopg2
    import psycopg2.extras as extras
    from dotenv import load_dotenv
    import os

    load_dotenv()
    user = os.getenv('USER_2')
    name = os.getenv('NAME_2')
    password = os.getenv('PASSWORD_2')
    host = os.getenv('HOST_2')
    port = os.getenv('PORT_2')

    CONNECTION_TIMEOUT = 10
    QUERY_TIMEOUT = 5

    columns_definition = ",\n    ".join([f"{col} {dtype}" for col, dtype in columns_type.items()])
    unique_constraint = f"UNIQUE ({', '.join(key_columns)})"

    try:
        conn = psycopg2.connect(
            dbname=name,
            user=user,
            password=password,
            host=host,
            port=port,
            connect_timeout=CONNECTION_TIMEOUT
        )
        conn.autocommit = False
        print("Соединение с базой данных установлено.")
        cur = conn.cursor()

        cur.execute(f"SET statement_timeout = {QUERY_TIMEOUT * 100};")

        # Создание таблицы с уникальным ограничением
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                {columns_definition},
                {unique_constraint}
            );
        """)
        print(f"Таблица {table_name} создана или уже существует.")

        columns = list(df.columns)
        tuples = list(df.itertuples(index=False, name=None))

        # Формируем запрос для upsert
        insert_query = f"""
            INSERT INTO {table_name} ({', '.join(columns)})
            VALUES %s
            ON CONFLICT ({', '.join(key_columns)}) DO UPDATE SET
            {', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col not in key_columns])};
        """
        extras.execute_values(cur, insert_query, tuples)

        conn.commit()
        print("Данные успешно добавлены в БД.")

    except psycopg2.OperationalError as e:
        print(f"Ошибка подключения к БД: {e}")
    except psycopg2.DatabaseError as e:
        print(f"Ошибка при загрузке данных в БД: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()
            print("Соединение с базой данных закрыто.")


# Глобальный лок для создания таблицы
table_creation_lock = asyncio.Lock()

async def create_insert_table_db_async(df: pd.DataFrame, table_name: str, columns_type: dict, key_columns: tuple):
    load_dotenv()

    conn = None
    try:
        conn = await asyncpg.connect(
            user=os.getenv('USER_2'),
            password=os.getenv('PASSWORD_2'),
            database=os.getenv('NAME_2'),
            host=os.getenv('HOST_2'),
            port=os.getenv('PORT_2'),
            timeout=10
        )

        # Проверка типов данных
        valid_types = ['INTEGER', 'BIGINT', 'SMALLINT', 'NUMERIC', 'DATE', 'TIMESTAMP', 'BOOLEAN', 'TEXT', 'VARCHAR']
        for col, dtype in columns_type.items():
            if not dtype.strip():
                raise ValueError(f"Пустой тип данных для колонки {col}")
            base_type = dtype.split('(')[0].strip().upper()
            if base_type not in valid_types:
                raise ValueError(f"Недопустимый тип данных для колонки {col}: {dtype}")
            if '(' in dtype and base_type not in ['NUMERIC', 'VARCHAR']:
                raise ValueError(f"Скобки недопустимы для типа {base_type} в колонке {col}: {dtype}")

        # Проверка и добавление отсутствующих колонок
        missing_cols = set(columns_type.keys()) - set(df.columns)
        for col in missing_cols:
            df[col] = None
            logging.info(f"Добавлена отсутствующая колонка {col} с None")
        extra_cols = set(df.columns) - set(columns_type.keys())
        if extra_cols:
            logging.warning(f"Лишние колонки в DataFrame: {extra_cols}, они будут проигнорированы")

        # Формирование SQL для колонок
        columns_definition = ", ".join([f"{col} {dtype}" for col, dtype in columns_type.items()])
        # создает SQL-выражение для уникального ограничения (UNIQUE constraint) в таблице базы данных.
        unique_constraint = f"CONSTRAINT unique_{table_name} UNIQUE ({', '.join(key_columns)})" if key_columns else ""

        # Отладка: вывод SQL-запроса
        create_table_query = f"""
            DO $$
            BEGIN
                IF NOT EXISTS (SELECT FROM pg_tables WHERE tablename = '{table_name}') THEN
                    CREATE TABLE {table_name} (
                        {columns_definition}{', ' + unique_constraint if unique_constraint else ''}
                    );
                END IF;
            END$$;
        """



        # Безопасное создание таблицы
        async with table_creation_lock:
            await conn.execute(create_table_query)
        
        # Подготовка данных для вставки
        columns = list(columns_type.keys())  # Используем только колонки из columns_type
        records = [tuple(None if pd.isna(x) else x for x in row) for row in df[columns].to_records(index=False)]

        # Формирование UPSERT-запроса
        updates = ', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col not in key_columns])
        query = f"""
            INSERT INTO {table_name} ({', '.join(columns)})
            VALUES ({', '.join([f'${i+1}' for i in range(len(columns))])})
            ON CONFLICT ON CONSTRAINT unique_{table_name}
            DO UPDATE SET {updates}
        """
        
        await conn.executemany(query, records)
        logging.info(f"Успешно сохранено {len(df)} строк в {table_name}")
        
    except Exception as e:
        logging.error(f"Ошибка при работе с БД: {str(e)}")
        raise
    finally:
        if conn:
            await conn.close()


# Данные по повторным отгрузкам
semaphore = asyncio.Semaphore(10)

async def re_shipment_info_get(account, api_token, date: str):
    """Функция позволяет получать данные о сборочных задниях, требующих повторную отгрузку"""
    
    url = "https://marketplace-api.wildberries.ru/api/v3/supplies/orders/reshipment"
    headers = {"Authorization": api_token}
    
    async with semaphore:
        async with aiohttp.ClientSession(headers=headers) as session:
            retry_count = 0
            while retry_count < 5:
                try:
                    async with session.get(url) as response:
                        print(f"HTTP статус: {response.status} по ЛК {account}")
                        
                        if response.status == 400:
                            err = await response.json()
                            print(f"Ошибка 400 {account}: {err.get('message') or err}")
                            break  # Выходим из цикла, т.к. 400 обычно не исправляется повтором
                            
                        if response.status == 429:
                            print(f"429 Too Many Requests для {account} — ждём минуту")
                            retry_count += 1
                            await asyncio.sleep(60)
                            continue
                            
                        response.raise_for_status()
                        res_data = await response.json()
                        
                        # Добавляем account к каждому элементу
                        if isinstance(res_data, list):
                            for item in res_data:
                                # Добавляем account к основному объекту
                                item['account'] = account
                                
                                # Добавляем account к каждому заказу внутри orders
                                if 'orders' in item and isinstance(item['orders'], list):
                                    for order in item['orders']:
                                        order['account'] = account
                                        
                            return res_data
                        else:
                            # Если ответ не список, добавляем account к объекту
                            res_data['account'] = account
                            
                            # И к orders если они есть
                            if 'orders' in res_data and isinstance(res_data['orders'], list):
                                for order in res_data['orders']:
                                    order['date'] = date
                                    order['account'] = account
                                    
                            return res_data  # Возвращаем как список для единообразия
                            
                except aiohttp.ClientError as e:
                    print(f"Сетевая ошибка для {account}: {e}")
                    retry_count += 1
                    await asyncio.sleep(30)
            
            # Если дошли сюда - все попытки исчерпаны
            print(f"Не удалось получить данные для {account} после 5 попыток")
            return []
        
def batchify(data, batch_size):
    """
    Splits data into batches of a specified size.

    Parameters:
    - data: The list of items to be batched.
    - batch_size: The size of each batch.

    Returns:
    - A generator yielding batches of data.
    """
    for i in range(0, len(data), batch_size):
        yield data[i:i + batch_size]



def create_insert_table_db_sync(df: pd.DataFrame, table_name: str, columns_type: dict, key_columns: tuple):
    load_dotenv()
    
    user = os.getenv('USER_2')
    password = os.getenv('PASSWORD_2')
    database = os.getenv('NAME_2') 
    host = os.getenv('HOST_2')
    port = os.getenv('PORT_2')
    
    connection_string = f"postgresql://{user}:{password}@{host}:{port}/{database}"
    engine = None
    
    try:
        engine = create_engine(connection_string)
        
        # Проверка типов данных
        valid_types = ['INTEGER', 'BIGINT', 'SMALLINT', 'NUMERIC', 'DATE', 'TIMESTAMP', 'BOOLEAN', 'TEXT', 'VARCHAR']
        for col, dtype in columns_type.items():
            if not dtype.strip():
                raise ValueError(f"Пустой тип данных для колонки {col}")
            base_type = dtype.split('(')[0].strip().upper()
            if base_type not in valid_types:
                raise ValueError(f"Недопустимый тип данных для колонки {col}: {dtype}")

        # Проверка и добавление отсутствующих колонок
        missing_cols = set(columns_type.keys()) - set(df.columns)
        for col in missing_cols:
            df[col] = None
            logging.info(f"Добавлена отсутствующая колонка {col} с None")
        
        extra_cols = set(df.columns) - set(columns_type.keys())
        if extra_cols:
            logging.warning(f"Лишние колонки в DataFrame: {extra_cols}, они будут проигнорированы")

        # Формирование SQL для создания таблицы
        columns_definition = ", ".join([f"{col} {dtype}" for col, dtype in columns_type.items()])
        unique_constraint = f", CONSTRAINT unique_{table_name} UNIQUE ({', '.join(key_columns)})" if key_columns else ""

        create_table_query = f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                {columns_definition}{unique_constraint}
            )
        """

        # Создание таблицы
        with engine.connect() as conn:
            conn.execute(text(create_table_query))
            conn.commit()
        
        # Подготовка данных для вставки
        columns = list(columns_type.keys())
        
        # Создание временной таблицы для UPSERT
        temp_table = f"temp_{table_name}"
        
        # Вставляем данные во временную таблицу
        df[columns].to_sql(temp_table, engine, if_exists='replace', index=False)
        
        # Выполняем UPSERT из временной таблицы с явным преобразованием типов
        with engine.connect() as conn:
            if key_columns:
                # Формируем SELECT с явным преобразованием типов
                select_columns = []
                for col in columns:
                    if columns_type[col].upper() == 'TIMESTAMP':
                        select_columns.append(f"CAST({col} AS TIMESTAMP)")  # Явное преобразование
                    else:
                        select_columns.append(col)
                
                updates = ', '.join([f"{col}=EXCLUDED.{col}" for col in columns if col not in key_columns])
                upsert_query = f"""
                    INSERT INTO {table_name} ({', '.join(columns)})
                    SELECT {', '.join(select_columns)} FROM {temp_table}
                    ON CONFLICT ({', '.join(key_columns)}) 
                    DO UPDATE SET {updates}
                """
            else:
                # Простая вставка с преобразованием типов
                select_columns = []
                for col in columns:
                    if columns_type[col].upper() == 'TIMESTAMP':
                        select_columns.append(f"CAST({col} AS TIMESTAMP)")
                    else:
                        select_columns.append(col)
                
                upsert_query = f"""
                    INSERT INTO {table_name} ({', '.join(columns)})
                    SELECT {', '.join(select_columns)} FROM {temp_table}
                """
            
            conn.execute(text(upsert_query))
            # Удаляем временную таблицу
            conn.execute(text(f"DROP TABLE {temp_table}"))
            conn.commit()
        
        logging.info(f"Успешно сохранено {len(df)} строк в {table_name}")
        
    except SQLAlchemyError as e:
        logging.error(f"Ошибка при работе с БД: {str(e)}")
        raise
    except Exception as e:
        logging.error(f"Неожиданная ошибка: {str(e)}")
        raise
    finally:
        if engine:
            engine.dispose()

# Подключение к базе данных
def create_connection(db_name, db_user, db_password, db_host, db_port):
    connection = None
    try:
        connection = psycopg2.connect(
            database=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port,
        )
        print(f"Соединение с БД PostgreSQL успешно установлено в {datetime.now().strftime('%Y-%m-%d')}")
    except OperationalError as error:
        print(f"Произошла ошибка при подключении к БД PostgreSQL {error}")
    return connection